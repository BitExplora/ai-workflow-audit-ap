"""Builds reports/value_model.xlsx from sandbox results. Formulas, not hardcodes."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
res = json.load(open(os.path.join(HERE, "..", "harness", "results.json")))

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial", color="000000")
BOLD = Font(name="Arial", bold=True)
WHITEBOLD = Font(name="Arial", bold=True, color="FFFFFF")
HDR = PatternFill("solid", fgColor="1F4E78")
YEL = PatternFill("solid", fgColor="FFFF00")
GREY = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CUR = '$#,##0;($#,##0);-'
PCT = '0.0%'

wb = Workbook()

# ---------------- Assumptions ----------------
a = wb.active
a.title = "Assumptions"
a.column_dimensions["A"].width = 46
a.column_dimensions["B"].width = 16
a.column_dimensions["C"].width = 40
a["A1"] = "AP Three-Way Match — Audit Assumptions"; a["A1"].font = BOLD

def row(ws, r, label, val, fmt=None, blue=True, note="", yellow=False):
    ws[f"A{r}"] = label; ws[f"A{r}"].font = BLACK
    c = ws[f"B{r}"]; c.value = val
    c.font = BLUE if blue else BLACK
    if fmt: c.number_format = fmt
    if yellow: c.fill = YEL
    if note: ws[f"C{r}"] = note; ws[f"C{r}"].font = Font(name="Arial", italic=True, size=9, color="808080")

a["A3"] = "Operational"; a["A3"].font = BOLD; a["A3"].fill = GREY
row(a, 4, "Annual invoice volume", 120000, '#,##0', note="Mid-market manufacturer (~10k/mo)")
row(a, 5, "Avg manual handle time (min/invoice)", 9, '0.0', note="Discovery interview input")
row(a, 6, "Loaded AP cost ($/hour)", 35, CUR, note="From client FP&A")
row(a, 7, "Loaded AP cost ($/min)", "=B6/60", CUR, blue=False)
row(a, 8, "Clean / auto-payable share", res["clean_share"], PCT, note="Sandbox: harness/results.json", yellow=True)
row(a, 9, "Exception share", "=1-B8", PCT, blue=False)
row(a, 10, "Hybrid review time (min/flagged item)", 4, '0.0')
row(a, 11, "Spot-QA time on clean (min/invoice)", 0.5, '0.0')

a["A13"] = "Risk"; a["A13"].font = BOLD; a["A13"].fill = GREY
row(a, 14, "Fraud rate (% of volume)", round(res["fraud_support"]/res["total_invoices"], 4), PCT, note="Sandbox label rate")
row(a, 15, "Residual fraud miss rate if FULLY automated", round(res["fraud_missed"]/res["fraud_support"], 4), PCT, note="Sandbox fraud recall gap", yellow=True)
row(a, 16, "Avg net loss per missed fraud ($)", 4500, CUR, note="Client risk team estimate")
row(a, 17, "Duplicate rate (% of volume)", round(res["duplicate_support"]/res["total_invoices"], 4), PCT)
row(a, 18, "Manual duplicate miss rate (today)", 0.30, PCT, note="Humans miss dupes in manual flow")
row(a, 19, "Net loss per missed duplicate ($)", 900, CUR)

a["A21"] = "Upside"; a["A21"].font = BOLD; a["A21"].fill = GREY
row(a, 22, "Avg invoice value ($)", 1800, CUR)
row(a, 23, "Early-pay discount available (%)", 0.02, PCT)
row(a, 24, "Current discount capture rate", 0.40, PCT)
row(a, 25, "Target discount capture rate", 0.85, PCT)

a["A27"] = "Cost to run (annual)"; a["A27"].font = BOLD; a["A27"].fill = GREY
row(a, 28, "Software / platform", 60000, CUR)
row(a, 29, "Implementation (amortized)", 40000, CUR)
row(a, 30, "Oversight / governance", 50000, CUR)

# ---------------- Workflow Audit (matrix) ----------------
w = wb.create_sheet("Workflow Audit")
for col, wdt in zip("ABCDE", [34, 10, 10, 16, 50]):
    w.column_dimensions[col].width = wdt
w["A1"] = "Sub-workflow scoring — impact x risk -> decision"; w["A1"].font = BOLD
heads = ["Sub-workflow", "Impact", "Risk", "Decision", "Rationale"]
for i, h in enumerate(heads):
    c = w.cell(row=3, column=i+1, value=h); c.font = WHITEBOLD; c.fill = HDR; c.border = BORD; c.alignment = Alignment(horizontal="center")

# impact, risk (1-5 judgments = blue inputs), rationale; decision = formula from matrix
flows = [
    ("Invoice data capture (OCR/extract)", 4, 2, "Deterministic; reversible before posting"),
    ("Three-way match (clean invoices)", 5, 2, "81% of volume; high accuracy in sandbox"),
    ("Duplicate detection / auto-hold", 4, 2, "100% recall; catches dupes humans miss"),
    ("Price / quantity mismatch resolution", 4, 3, "Needs judgment on tolerance + vendor comms"),
    ("Missing-PO / non-PO routing", 3, 3, "Policy-dependent; route to owner"),
    ("Tax / GL coding", 3, 3, "Automatable with periodic sampling"),
    ("Fraud / bank-change screening", 5, 5, "Sandbox missed 50% of fraud; irreversible"),
    ("Payment approval & release", 5, 5, "Binding cash out; must gate above threshold"),
]
r0 = 4
for j, (name, imp, rsk, rat) in enumerate(flows):
    r = r0 + j
    w.cell(row=r, column=1, value=name).font = BLACK
    ci = w.cell(row=r, column=2, value=imp); ci.font = BLUE; ci.alignment = Alignment(horizontal="center")
    cr = w.cell(row=r, column=3, value=rsk); cr.font = BLUE; cr.alignment = Alignment(horizontal="center")
    # Matrix rule: low impact -> Manual/Kill; high impact & high risk -> Hybrid;
    # high impact & low risk -> Automate; else Hybrid
    dec = (f'=IF(B{r}<3,"Manual/Kill",'
           f'IF(AND(B{r}>=4,C{r}>=4),"Hybrid",'
           f'IF(AND(B{r}>=4,C{r}<=3),"Automate","Hybrid")))')
    cd = w.cell(row=r, column=4, value=dec); cd.font = BLACK; cd.alignment = Alignment(horizontal="center")
    w.cell(row=r, column=5, value=rat).font = Font(name="Arial", size=9)
    for col in range(1, 6):
        w.cell(row=r, column=col).border = BORD

# ---------------- Value Model ----------------
v = wb.create_sheet("Value Model")
v.column_dimensions["A"].width = 52
for col in "BC":
    v.column_dimensions[col].width = 18
v["A1"] = "Annual value — target operating model vs naive full automation"; v["A1"].font = BOLD
v["B3"] = "Recommended (Hybrid)"; v["C3"] = "Naive full automation"
for c in ("B3", "C3"): v[c].font = WHITEBOLD; v[c].fill = HDR; v[c].alignment = Alignment(horizontal="center")

A = "Assumptions"
def vrow(r, label, bform, cform=None, fmt=CUR, bold=False):
    v[f"A{r}"] = label; v[f"A{r}"].font = BOLD if bold else BLACK
    v[f"B{r}"] = bform; v[f"B{r}"].number_format = fmt; v[f"B{r}"].font = BLACK
    if cform is not None:
        v[f"C{r}"] = cform; v[f"C{r}"].number_format = fmt; v[f"C{r}"].font = BLACK

vrow(4, "Current manual labor cost", f"='{A}'!B4*'{A}'!B5*'{A}'!B7", f"='{A}'!B4*'{A}'!B5*'{A}'!B7")
# target labor: hybrid reviews exceptions; full-auto reviews nothing
vrow(5, "Target labor cost",
     f"=('{A}'!B4*'{A}'!B8*'{A}'!B11+'{A}'!B4*'{A}'!B9*'{A}'!B10)*'{A}'!B7",
     f"='{A}'!B4*'{A}'!B8*'{A}'!B11*'{A}'!B7")
vrow(6, "Labor saved", "=B4-B5", "=C4-C5", bold=True)
vrow(7, "Duplicate losses prevented",
     f"='{A}'!B4*'{A}'!B17*'{A}'!B18*'{A}'!B19",
     f"='{A}'!B4*'{A}'!B17*'{A}'!B18*'{A}'!B19")
vrow(8, "Early-pay discount capture gain",
     f"='{A}'!B4*'{A}'!B22*'{A}'!B23*('{A}'!B25-'{A}'!B24)",
     f"='{A}'!B4*'{A}'!B22*'{A}'!B23*('{A}'!B25-'{A}'!B24)")
vrow(9, "Fraud losses incurred (cost)",
     "=0",
     f"=-'{A}'!B4*'{A}'!B14*'{A}'!B15*'{A}'!B16")
v["A9"].font = BLACK
vrow(10, "Cost to run", f"=-SUM('{A}'!B28:B30)", f"=-SUM('{A}'!B28:B30)")
vrow(11, "Net annual value", "=B6+B7+B8+B9+B10", "=C6+C7+C8+C9+C10", bold=True)
for c in ("A11", "B11", "C11"):
    v[c].fill = YEL
vrow(13, "Advantage of hybrid over full automation", "=B11-C11", None, bold=True)

wb.save(os.path.join(HERE, "value_model.xlsx"))
print("saved value_model.xlsx")
